import streamlit as st
import fitz  # PyMuPDF
from docx import Document
from groq import Groq
import pandas as pd
import io
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

# --- 1. CONFIGURATION ---

GROQ_API_KEY = "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"
client = Groq(api_key=GROQ_API_KEY)
MODEL_ID = "llama-3.3-70b-versatile"

# --- 2. PAGE CONFIG ---
st.set_page_config(page_title="Muhali QA | Test Case Designer", page_icon=None, layout="wide")

# --- 3. CUSTOM CSS ---
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Mono:wght@400;500&display=swap');

/* ---- GLOBAL RESET ---- */
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

html, body, .stApp {
    background-color: #0D0F12 !important;
    color: #E8E8E8 !important;
    font-family: 'DM Sans', sans-serif !important;
}

/* Hide Streamlit default chrome */
#MainMenu, footer, header { visibility: hidden; }
.block-container {
    padding: 2.5rem 3rem 4rem 3rem !important;
    max-width: 1200px !important;
}

/* ---- SIDEBAR ---- */
[data-testid="stSidebar"] {
    background-color: #111318 !important;
    border-right: 1px solid #1E2028 !important;
}
[data-testid="stSidebar"] * {
    color: #A0A4B0 !important;
    font-family: 'DM Sans', sans-serif !important;
}
[data-testid="stSidebar"] .sidebar-brand {
    font-size: 1.1rem;
    font-weight: 600;
    color: #FFFFFF !important;
    letter-spacing: 0.04em;
    text-transform: uppercase;
}
[data-testid="stSidebar"] hr {
    border-color: #1E2028 !important;
    margin: 1.2rem 0 !important;
}

/* ---- PAGE HEADER ---- */
.page-header {
    border-bottom: 1px solid #1E2028;
    padding-bottom: 2rem;
    margin-bottom: 2.5rem;
}
.page-label {
    font-family: 'DM Mono', monospace;
    font-size: 0.72rem;
    letter-spacing: 0.18em;
    text-transform: uppercase;
    color: #4A90D9;
    margin-bottom: 0.6rem;
}
.page-title {
    font-size: 2.2rem;
    font-weight: 300;
    color: #F0F0F0;
    line-height: 1.2;
    letter-spacing: -0.02em;
}
.page-title strong {
    font-weight: 600;
    color: #FFFFFF;
}
.page-subtitle {
    font-size: 0.95rem;
    color: #606470;
    margin-top: 0.5rem;
    font-weight: 400;
}

/* ---- UPLOAD AREA ---- */
[data-testid="stFileUploader"] {
    background-color: #111318 !important;
    border: 1px dashed #2A2D38 !important;
    border-radius: 10px !important;
    padding: 1.5rem !important;
    transition: border-color 0.2s ease;
}
[data-testid="stFileUploader"]:hover {
    border-color: #4A90D9 !important;
}
[data-testid="stFileUploader"] label {
    color: #606470 !important;
    font-size: 0.9rem !important;
}
[data-testid="stFileUploadDropzone"] {
    background: transparent !important;
}

/* ---- SECTION LABEL ---- */
.section-label {
    font-family: 'DM Mono', monospace;
    font-size: 0.68rem;
    letter-spacing: 0.16em;
    text-transform: uppercase;
    color: #404450;
    margin-bottom: 0.8rem;
}

/* ---- FILE PILL ---- */
.file-pill {
    display: inline-flex;
    align-items: center;
    gap: 0.5rem;
    background-color: #161920;
    border: 1px solid #2A2D38;
    border-radius: 6px;
    padding: 0.5rem 1rem;
    font-size: 0.88rem;
    color: #C0C4D0;
    font-family: 'DM Mono', monospace;
}
.file-dot {
    width: 7px;
    height: 7px;
    border-radius: 50%;
    background-color: #3CB97A;
    display: inline-block;
}

/* ---- GENERATE BUTTON ---- */
.stButton > button {
    width: 100% !important;
    background-color: #4A90D9 !important;
    color: #FFFFFF !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 500 !important;
    font-size: 0.9rem !important;
    letter-spacing: 0.04em !important;
    border: none !important;
    border-radius: 8px !important;
    height: 3rem !important;
    transition: background-color 0.2s ease, transform 0.1s ease !important;
    box-shadow: 0 0 0 0 rgba(74, 144, 217, 0) !important;
}
.stButton > button:hover {
    background-color: #3A7EC8 !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 20px rgba(74, 144, 217, 0.25) !important;
}
.stButton > button:active {
    transform: translateY(0px) !important;
}

/* ---- DIVIDER ---- */
hr {
    border: none !important;
    border-top: 1px solid #1E2028 !important;
    margin: 2rem 0 !important;
}

/* ---- PREVIEW TABLE ---- */
.preview-wrapper {
    background-color: #111318;
    border: 1px solid #1E2028;
    border-radius: 10px;
    padding: 1.5rem 2rem;
    margin-top: 0.5rem;
}
.stMarkdown table {
    width: 100% !important;
    border-collapse: collapse !important;
    font-size: 0.875rem !important;
    font-family: 'DM Sans', sans-serif !important;
}
.stMarkdown table th {
    background-color: #161920 !important;
    color: #4A90D9 !important;
    font-weight: 600 !important;
    font-size: 0.75rem !important;
    letter-spacing: 0.1em !important;
    text-transform: uppercase !important;
    padding: 0.75rem 1rem !important;
    border: 1px solid #1E2028 !important;
    text-align: left !important;
}
.stMarkdown table td {
    padding: 0.65rem 1rem !important;
    border: 1px solid #1A1D24 !important;
    color: #C0C4D0 !important;
    vertical-align: top !important;
    line-height: 1.5 !important;
}
.stMarkdown table tr:nth-child(even) td {
    background-color: #0F1115 !important;
}
.stMarkdown table tr:hover td {
    background-color: #161C26 !important;
}

/* ---- DOWNLOAD BUTTON ---- */
[data-testid="stDownloadButton"] > button {
    background-color: transparent !important;
    color: #4A90D9 !important;
    border: 1px solid #2A3A50 !important;
    border-radius: 8px !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 500 !important;
    font-size: 0.88rem !important;
    height: 3rem !important;
    letter-spacing: 0.03em !important;
    transition: all 0.2s ease !important;
}
[data-testid="stDownloadButton"] > button:hover {
    background-color: #4A90D9 !important;
    color: #FFFFFF !important;
    border-color: #4A90D9 !important;
}

/* ---- ALERTS ---- */
[data-testid="stAlert"] {
    background-color: #111318 !important;
    border: 1px solid #2A2D38 !important;
    border-radius: 8px !important;
    color: #A0A4B0 !important;
    font-size: 0.88rem !important;
}

/* ---- SPINNER ---- */
[data-testid="stSpinner"] {
    color: #4A90D9 !important;
}

/* ---- CAPTION / FOOTER ---- */
.footer-bar {
    margin-top: 4rem;
    padding-top: 1.5rem;
    border-top: 1px solid #1A1D24;
    display: flex;
    justify-content: space-between;
    align-items: center;
}
.footer-text {
    font-size: 0.78rem;
    color: #2E3240;
    font-family: 'DM Mono', monospace;
    letter-spacing: 0.06em;
}
.status-badge {
    display: inline-flex;
    align-items: center;
    gap: 0.4rem;
    font-size: 0.75rem;
    color: #3CB97A;
    font-family: 'DM Mono', monospace;
}
.status-dot {
    width: 6px;
    height: 6px;
    border-radius: 50%;
    background-color: #3CB97A;
}
</style>
""", unsafe_allow_html=True)


# --- 4. HELPERS ---
def extract_text(uploaded_file):
    ext = uploaded_file.name.split('.')[-1].lower()
    try:
        if ext == "pdf":
            uploaded_file.seek(0)
            with fitz.open(stream=uploaded_file.read(), filetype="pdf") as doc:
                return "".join([page.get_text() for page in doc])
        elif ext == "docx":
            doc = Document(uploaded_file)
            return "\n".join([p.text for p in doc.paragraphs])
        elif ext == "txt":
            return uploaded_file.read().decode("utf-8")
        return None
    except Exception as e:
        st.error(f"Extraction error: {e}")
        return None


def call_groq(raw_text: str) -> str:
    prompt = f"""You are a Senior QA Engineer at Muhali Software Solutions.
Your task is to convert the provided requirements document into a high-level TEST CASE Design.

Format your entire response as a Markdown table with EXACTLY these columns:
Test ID | Title | Expected Result | Priority | Status

RULES:
- Every entry in the Title column MUST start with the words 'Verify that'
- The Status column must ALWAYS contain only the word 'Pending' for every row.
- Provide at least 10 varied and meaningful test scenarios covering happy paths, edge cases, and negative tests.
- Do NOT include a 'Steps' or 'Actual Result' column.
- Do NOT add any text before or after the table — output the table only.
- Priority values must be one of: High, Medium, or Low.

DOCUMENT CONTENT:
{raw_text}
"""
    response = client.chat.completions.create(
        model=MODEL_ID,
        messages=[
            {"role": "system", "content": "You are a Senior QA Engineer. You only respond with clean Markdown tables. No extra text."},
            {"role": "user", "content": prompt}
        ],
        max_tokens=4096,
        temperature=0.3
    )
    return response.choices[0].message.content


def convert_to_excel(raw_ai_text: str):
    lines = raw_ai_text.strip().split('\n')
    data = []
    for line in lines:
        if "|" in line and "---" not in line:
            cols = [c.strip() for c in line.split('|') if c.strip()]
            if len(cols) >= 4:
                while len(cols) < 5:
                    cols.append("Pending")
                data.append(cols[:5])

    if not data or len(data) < 2:
        return None

    df = pd.DataFrame(data[1:], columns=data[0])
    output = io.BytesIO()
    sheet_name = 'QA Tracker'

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        wrap_align = Alignment(wrap_text=True, vertical='top', horizontal='left')
        header_fill = PatternFill(start_color="1A2B3C", end_color="1A2B3C", fill_type="solid")
        header_font = Font(bold=True, color="4A90D9", name="Calibri")

        dv = DataValidation(type="list", formula1='"Pass,Fail,Pending"', allow_blank=True)
        worksheet.add_data_validation(dv)

        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = wrap_align
                if cell.row == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    if cell.column == 5:
                        dv.add(cell)

        worksheet.column_dimensions['A'].width = 12
        worksheet.column_dimensions['B'].width = 35
        worksheet.column_dimensions['C'].width = 60
        worksheet.column_dimensions['D'].width = 12
        worksheet.column_dimensions['E'].width = 15

    return output.getvalue()


# --- 5. SIDEBAR ---
with st.sidebar:
    st.markdown('<div class="sidebar-brand">Muhali Framework</div>', unsafe_allow_html=True)
    st.markdown("---")
    st.markdown('<div style="font-size:0.82rem; color:#606470; line-height:1.7;">Agency<br><span style="color:#A0A4B0; font-weight:500;">Muhali Software Solutions</span></div>', unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div style="font-size:0.82rem; color:#606470; line-height:1.7;">AI Engine<br><span style="color:#A0A4B0; font-weight:500;">Meta Llama 3.3 via Groq</span></div>', unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div style="font-size:0.82rem; color:#606470; line-height:1.7;">Accepted Formats<br><span style="color:#A0A4B0; font-weight:500;">PDF, DOCX, TXT</span></div>', unsafe_allow_html=True)
    st.markdown("---")
    st.markdown('<div style="font-size:0.8rem; color:#404450; line-height:1.6;">Upload a BRS or BRD document to generate a structured, ready-to-execute QA test case tracker.</div>', unsafe_allow_html=True)


# --- 6. MAIN CONTENT ---

# Header
st.markdown("""
<div class="page-header">
    <div class="page-label">QA Automation Suite</div>
    <div class="page-title">Test Case <strong>Designer</strong></div>
    <div class="page-subtitle">Convert requirement documents into structured, executable test case designs</div>
</div>
""", unsafe_allow_html=True)

# Upload section
st.markdown('<div class="section-label">Step 01 — Upload Document</div>', unsafe_allow_html=True)
uploaded_file = st.file_uploader("Upload Requirement Document", type=["pdf", "docx", "txt"], label_visibility="collapsed")

if uploaded_file:
    st.markdown("<br>", unsafe_allow_html=True)

    col1, col2 = st.columns([3, 1], gap="medium")
    with col1:
        st.markdown(f"""
        <div class="file-pill">
            <span class="file-dot"></span>
            {uploaded_file.name}
        </div>
        """, unsafe_allow_html=True)
    with col2:
        generate_btn = st.button("Generate Test Cases")

    if generate_btn:
        st.markdown("<br>", unsafe_allow_html=True)
        with st.spinner("Analysing document and generating test cases..."):
            raw_text = extract_text(uploaded_file)
            if raw_text:
                try:
                    summary_text = call_groq(raw_text)

                    st.markdown("---")
                    st.markdown('<div class="section-label">Step 02 — Preview</div>', unsafe_allow_html=True)
                    st.markdown('<div class="preview-wrapper">', unsafe_allow_html=True)
                    st.markdown(summary_text)
                    st.markdown('</div>', unsafe_allow_html=True)

                    excel_data = convert_to_excel(summary_text)

                    if excel_data:
                        st.markdown("---")
                        st.markdown('<div class="section-label">Step 03 — Export</div>', unsafe_allow_html=True)
                        st.download_button(
                            label="Download Execution Tracker (.xlsx)",
                            data=excel_data,
                            file_name=f"QA_Tracker_{uploaded_file.name.split('.')[0]}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    else:
                        st.error("Could not parse the AI response into a table. Please try again.")

                except Exception as e:
                    st.error(f"Error: {e}")

# Footer
st.markdown("""
<div class="footer-bar">
    <div class="footer-text">MUHALI SOFTWARE SOLUTIONS / 2026</div>
    <div class="status-badge">
        <span class="status-dot"></span>
        MUHALI SOFTWARE SOLUTIONS
    </div>
</div>
""", unsafe_allow_html=True)
